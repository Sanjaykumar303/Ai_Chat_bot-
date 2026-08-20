# services/xlsx_export.py - takes rows already fetched by the existing
# DB pipeline (never queries anything itself), produces real, openable
# .xlsx bytes. No Gemini/DB calls anywhere in this module, so nothing
# here needs monkeypatching.

import datetime
import io
from decimal import Decimal

from openpyxl import load_workbook

from services import xlsx_export


def _load(file_bytes):
    return load_workbook(io.BytesIO(file_bytes))


def test_build_xlsx_writes_headers_and_rows_from_real_dicts():
    rows = [
        {"account": "Sales", "amount": 1000},
        {"account": "Refunds", "amount": -50},
    ]

    workbook = _load(xlsx_export.build_xlsx("Income", rows))
    sheet = workbook.active

    assert sheet.title == "Income"
    assert [cell.value for cell in sheet[1]] == ["account", "amount"]
    assert [cell.value for cell in sheet[2]] == ["Sales", 1000]
    assert [cell.value for cell in sheet[3]] == ["Refunds", -50]


def test_build_xlsx_never_fabricates_rows_that_were_not_returned():
    # Exactly the two rows that were passed in - nothing padded, nothing
    # invented, nothing dropped.
    rows = [{"id": 1}, {"id": 2}, {"id": 3}]

    sheet = _load(xlsx_export.build_xlsx("Records", rows)).active

    data_row_count = sum(1 for _ in sheet.iter_rows(min_row=2))
    assert data_row_count == len(rows)


def test_build_xlsx_on_empty_rows_says_so_honestly_rather_than_fabricating():
    sheet = _load(xlsx_export.build_xlsx("Income", [])).active

    assert sheet["A1"].value == "No matching records were found."
    assert sheet.max_row == 1


def test_build_xlsx_converts_decimal_money_columns_to_plain_numbers():
    rows = [{"total": Decimal("1234.56")}]

    sheet = _load(xlsx_export.build_xlsx("Income", rows)).active

    assert sheet["A2"].value == 1234.56
    assert isinstance(sheet["A2"].value, float)


def test_build_xlsx_passes_through_dates_and_none_natively():
    rows = [{"txn_date": datetime.date(2026, 8, 1), "note": None}]

    sheet = _load(xlsx_export.build_xlsx("Income", rows)).active

    # openpyxl itself normalizes a written date() to a datetime() on
    # read-back (there's no distinct "date-only" cell type at the XML
    # level) - comparing just the date portion is what actually confirms
    # the value round-tripped correctly, not a stricter type match
    # openpyxl's own format doesn't support anyway.
    assert sheet["A2"].value.date() == datetime.date(2026, 8, 1)
    assert sheet["B2"].value is None


def test_build_xlsx_header_row_is_bold():
    sheet = _load(xlsx_export.build_xlsx("Income", [{"account": "Sales"}])).active
    assert sheet["A1"].font.bold is True


def test_build_xlsx_truncates_sheet_title_to_excels_own_limit():
    long_title = "x" * 50
    sheet = _load(xlsx_export.build_xlsx(long_title, [{"id": 1}])).active
    assert len(sheet.title) <= 31
